import rawpy
import cv2
import numpy as np
import os
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Tuple, Optional
from multiprocessing import Pool, cpu_count
import torch
from segment_anything import sam_model_registry, SamPredictor
import clip
from PIL import Image
import pandas as pd
import openpyxl
import argparse
import json
import yaml
import logging
from logging.handlers import RotatingFileHandler
from utils import get_all_subdirectories, load_processed_directories, save_processed_directory


class PhotoPicker:
    def __init__(
        self,
        config_path: str,
        debug: bool = False,
        count_only: bool = False,
        force_reprocess: bool = False,
        log_level: str = "INFO",
    ):
        """
        Initialize the PhotoJudge with configuration and options.

        Args:
            config_path: Path to config file containing root directory
            debug: If True, only process up to 5 photos per daily directory
            count_only: If True, only count photos without processing
            force_reprocess: If True, process all directories even if previously processed
            log_level: Logging level (DEBUG, INFO, WARNING, ERROR)
        """
        self.config_path = config_path
        self.debug = debug
        self.count_only = count_only
        self.force_reprocess = force_reprocess
        self.processed_file = "processed_directories.json"

        # Initialize logger
        self.logger = self._setup_logging(log_level)

        # Will be initialized when needed
        self.models = None
        self.processed_dirs = None

    def _setup_logging(self, log_level: str = "INFO") -> logging.Logger:
        """Configure logging for the application"""
        if not os.path.exists("logs"):
            os.makedirs("logs")

        logger = logging.getLogger(__name__)
        logger.setLevel(getattr(logging, log_level))

        # Create formatters
        file_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

        # Create and configure file handler
        file_handler = RotatingFileHandler("logs/photo_judge.log", maxBytes=10 * 1024 * 1024, backupCount=5)
        file_handler.setLevel(getattr(logging, log_level))
        file_handler.setFormatter(file_formatter)

        # Add handler to logger
        logger.addHandler(file_handler)

        return logger

    def _load_models(self):
        """Initialize ML models if not already initialized"""
        if self.models is None:
            self.logger.info("Initializing ML models...")
            # Initialize SAM
            model_type = "vit_h"
            checkpoint = "sam_vit_h_4b8939.pth"
            device = "cuda" if torch.cuda.is_available() else "cpu"

            sam = sam_model_registry[model_type](checkpoint=checkpoint)
            sam.to(device=device)
            predictor = SamPredictor(sam)

            # Initialize CLIP
            clip_model, preprocess = clip.load("ViT-B/32", device=device)

            # Define airline prompts
            airline_prompts = [
                # US Carriers
                "a photo of an American Airlines aircraft",
                "a photo of a United Airlines aircraft",
                "a photo of a Delta Airlines aircraft",
                "a photo of a Southwest Airlines aircraft",
                "a photo of a JetBlue aircraft",
                "a photo of an Alaska Airlines aircraft",
                "a photo of a Spirit Airlines aircraft",
                "a photo of a Frontier Airlines aircraft",
                "a photo of a Hawaiian Airlines aircraft",
                "a photo of an Allegiant Air aircraft",
                # European Carriers
                "a photo of a Lufthansa aircraft",
                "a photo of a British Airways aircraft",
                "a photo of an Air France aircraft",
                "a photo of a KLM aircraft",
                "a photo of an Iberia aircraft",
                "a photo of a Turkish Airlines aircraft",
                "a photo of a Swiss International Airlines aircraft",
                "a photo of a Scandinavian Airlines aircraft",
                "a photo of an Austrian Airlines aircraft",
                "a photo of a TAP Air Portugal aircraft",
                "a photo of a LOT Polish Airlines aircraft",
                "a photo of a Finnair aircraft",
                "a photo of a Norwegian Air Shuttle aircraft",
                "a photo of an easyJet aircraft",
                "a photo of a Ryanair aircraft",
                "a photo of a Wizz Air aircraft",
                # Middle Eastern Carriers
                "a photo of an Emirates aircraft",
                "a photo of a Qatar Airways aircraft",
                "a photo of an Etihad Airways aircraft",
                # Asian Carriers
                "a photo of a Singapore Airlines aircraft",
                "a photo of a Cathay Pacific aircraft",
                "a photo of a Japan Airlines aircraft",
                "a photo of an All Nippon Airways aircraft",
                "a photo of a Korean Air aircraft",
                "a photo of an Asiana Airlines aircraft",
                "a photo of an Air China aircraft",
                "a photo of a China Eastern aircraft",
                "a photo of a China Southern aircraft",
                # Other Major Carriers
                "a photo of an Air Canada aircraft",
                "a photo of a Qantas aircraft",
                "a photo of a Virgin Atlantic aircraft",
                "a photo of a Virgin Australia aircraft",
                "a photo of an Air New Zealand aircraft",
                "a photo of a LATAM Airlines aircraft",
                "a photo of an Aeromexico aircraft",
                "a photo of a Copa Airlines aircraft",
                "a photo of an Avianca aircraft",
            ]

            # Pre-compute text features
            text_tokens = clip.tokenize(airline_prompts).to(device)
            with torch.no_grad():
                text_features = clip_model.encode_text(text_tokens)
                text_features /= text_features.norm(dim=-1, keepdim=True)

            self.models = (predictor, clip_model, preprocess, text_features, airline_prompts, device)
            self.logger.info("ML models initialized successfully")

    def _load_config(self) -> str:
        """Load root directory from config file"""
        if self.config_path.endswith(".json"):
            with open(self.config_path, "r") as f:
                config = json.load(f)
        elif self.config_path.endswith(".yaml") or self.config_path.endswith(".yml"):
            with open(self.config_path, "r") as f:
                config = yaml.safe_load(f)
        else:
            raise ValueError("Config file must be either JSON or YAML format")

        if isinstance(config, str):
            root_dir = config
        elif isinstance(config, dict) and "root_directory" in config:
            root_dir = config["root_directory"]
        else:
            raise ValueError("Config must be a string path or a dict with 'root_directory' key")

        if not os.path.isdir(root_dir):
            raise ValueError(f"Root directory '{root_dir}' does not exist")

        self.logger.info(f"Using root directory: {root_dir}")
        return root_dir

    def process_directory(self, directory_path: str) -> Optional[int]:
        """
        Process a single directory.
        Returns the number of photos found if in count-only mode, None otherwise.
        """
        raw_extensions = {".nef"}
        all_photo_paths = []

        try:
            current_photos = [
                os.path.join(directory_path, f)
                for f in os.listdir(directory_path)
                if os.path.splitext(f.lower())[1] in raw_extensions
            ]

            if not current_photos:
                self.logger.info(f"No photos found in {directory_path}")
                return 0 if self.count_only else None

            if self.debug:
                current_photos = current_photos[:5]
                self.logger.info(f"Debug mode: Processing only 5 photos")
            else:
                self.logger.info(f"Found {len(current_photos)} photos to process")

            all_photo_paths.extend(current_photos)

            if self.count_only:
                return len(all_photo_paths)

            # Process photos
            photos = []
            for photo_path in all_photo_paths:
                try:
                    assessment = self.assess_image(photo_path)
                    if assessment:
                        photos.append(assessment)
                except Exception as e:
                    self.logger.error(f"Error processing {photo_path}: {str(e)}")
                    continue

            if photos:
                # Group similar photos
                sequences = self.group_similar_photos(photos)

                # Rank photos in each sequence
                ranked_sequences = [self.rank_photos_in_sequence(seq) for seq in sequences]

                # Create or update assessment Excel file
                output_path = "photo_assessment.xlsx"
                self.create_assessment_excel(photos, ranked_sequences, output_path, directory_path)

                self.logger.info(
                    f"Successfully processed {len(photos)} photos and updated assessment file: {output_path}"
                )
            else:
                self.logger.warning(f"No photos were successfully processed in {directory_path}")

            return None

        except Exception as e:
            self.logger.error(f"Error processing directory {directory_path}: {str(e)}")
            return 0 if self.count_only else None

    def run(self) -> None:
        """Run the photo processing pipeline"""
        try:
            # Load configuration
            root_dir = self._load_config()

            # Load processed directories
            self.processed_dirs = load_processed_directories(self.processed_file, self.logger)

            # Initialize ML models if not in count-only mode
            if not self.count_only:
                self._load_models()

            total_photos = 0

            # Get all subdirectories with their processing status
            all_dirs = get_all_subdirectories(root_dir, self.processed_dirs, self.logger)

            self.logger.info(f"Found {len(all_dirs)} directories to process")

            # Process each directory
            for directory, is_processed in all_dirs:
                if not os.path.isdir(directory):
                    self.logger.warning(f"Warning: Directory '{directory}' does not exist, skipping...")
                    continue

                # Skip if already processed and not in debug mode and not forced
                if is_processed and not self.debug and not self.force_reprocess:
                    self.logger.info(f"\nSkipping already processed directory: {directory}")
                    self.logger.info(f"Last processed: {self.processed_dirs[directory]}")
                    continue

                self.logger.info(f"\nProcessing directory: {directory}")
                if self.count_only:
                    photos_count = self.process_directory(directory)
                    if photos_count is not None:
                        total_photos += photos_count
                else:
                    self.process_directory(directory)
                    # Save the processed directory at the daily level
                    save_processed_directory(directory, self.processed_file, self.logger)

            if self.count_only:
                self.logger.info(f"\nTotal photos found: {total_photos}")
            else:
                self.logger.info("\nAll directories processed successfully!")
        except ValueError as e:
            self.logger.error(f"Error: {str(e)}")
            return

    def _identify_airline(self, image):
        """Identify airline in the image using CLIP"""
        # Convert to PIL Image
        if isinstance(image, np.ndarray):
            image = Image.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))

        # Preprocess image
        image_input = self.models[2](image).unsqueeze(0).to(self.models[5])  # preprocess and device

        # Get image features
        with torch.no_grad():
            image_features = self.models[1].encode_image(image_input)  # clip_model
            image_features /= image_features.norm(dim=-1, keepdim=True)

        # Calculate similarity
        similarity = (100.0 * image_features @ self.models[3].T).softmax(dim=-1)  # text_features
        values, indices = similarity[0].topk(3)

        # Get top 3 matches with confidence scores
        matches = [
            (self.models[4][idx].split("of ")[-1].split(" aircraft")[0], float(val))  # airline_prompts
            for val, idx in zip(values, indices)
        ]

        return matches

    def _detect_sky(self, image, subject_mask=None):
        """
        Detect sky using multiple methods and combine results, excluding the subject area
        Returns: (sky_percentage, sky_mask)
        """
        h, w = image.shape[:2]

        # Create a mask excluding the subject area
        if subject_mask is not None:
            # Ensure subject_mask is binary
            subject_mask = (subject_mask > 0).astype(np.uint8)
            # Invert the subject mask to get background mask
            background_mask = 1 - subject_mask
        else:
            background_mask = np.ones((h, w), dtype=np.uint8)

        # 1. HSV Color-based detection
        hsv = cv2.cvtColor(image, cv2.COLOR_RGB2HSV)
        # Sky typically has high value (brightness) and low saturation
        lower_sky = np.array([90, 10, 180])  # Light blue to white
        upper_sky = np.array([130, 50, 255])
        sky_mask_hsv = cv2.inRange(hsv, lower_sky, upper_sky)
        # Apply background mask
        sky_mask_hsv = cv2.bitwise_and(sky_mask_hsv, sky_mask_hsv, mask=background_mask)

        # 2. RGB ratio method
        # Sky typically has higher blue values
        b, g, r = cv2.split(image)
        blue_ratio = b.astype(float) / (r.astype(float) + g.astype(float) + 1e-6)
        sky_mask_rgb = (blue_ratio > 0.4).astype(np.uint8) * 255
        # Apply background mask
        sky_mask_rgb = cv2.bitwise_and(sky_mask_rgb, sky_mask_rgb, mask=background_mask)

        # 3. Position-based weighting
        # Create a weight mask that favors upper portions of the image
        y_weights = np.linspace(1, 0, h)[:, np.newaxis]
        y_weights = np.tile(y_weights, (1, w))
        position_mask = (y_weights > 0.5).astype(np.uint8) * 255
        # Apply background mask
        position_mask = cv2.bitwise_and(position_mask, position_mask, mask=background_mask)

        # 4. Gradient-based detection
        gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
        sobelx = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
        sobely = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)
        gradient_magnitude = np.sqrt(sobelx**2 + sobely**2)
        # Sky typically has low gradient magnitude
        sky_mask_gradient = (gradient_magnitude < 30).astype(np.uint8) * 255
        # Apply background mask
        sky_mask_gradient = cv2.bitwise_and(sky_mask_gradient, sky_mask_gradient, mask=background_mask)

        # Combine all masks with weights
        combined_mask = (
            0.4 * sky_mask_hsv + 0.3 * sky_mask_rgb + 0.2 * position_mask + 0.1 * sky_mask_gradient
        ).astype(np.uint8)

        # Apply threshold to get final mask
        _, final_mask = cv2.threshold(combined_mask, 127, 255, cv2.THRESH_BINARY)

        # Calculate percentage of sky-like pixels in the background area
        if subject_mask is not None:
            # Calculate percentage only in the background area
            background_pixels = np.sum(background_mask)
            if background_pixels > 0:
                sky_percentage = np.sum(final_mask > 0) / background_pixels
            else:
                sky_percentage = 0
        else:
            # Calculate percentage of entire image
            sky_percentage = np.sum(final_mask > 0) / (h * w)

        return sky_percentage, final_mask

    def _get_subject_mask(self, image):
        """Get mask for the main subject using SAM with improved targeting"""
        # Convert image to RGB if it's not already
        if len(image.shape) == 2:
            image = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)

        # Set image in predictor
        self.models[0].set_image(image)  # predictor

        # Get image dimensions
        h, w = image.shape[:2]

        # Create a more targeted point grid focusing on the center region
        # This assumes the main subject is typically in the center of the frame
        center_x, center_y = w // 2, h // 2
        grid_size = 50  # Smaller grid size for more precise targeting

        # Create points in a grid around the center
        points = []
        for y in range(center_y - 200, center_y + 201, grid_size):
            for x in range(center_x - 200, center_x + 201, grid_size):
                if 0 <= x < w and 0 <= y < h:
                    points.append([x, y])

        # Add some points in the corners to help with negative examples
        corner_points = [[50, 50], [w - 50, 50], [50, h - 50], [w - 50, h - 50]]  # Top corners  # Bottom corners

        # Combine center points and corner points
        points = np.array(points + corner_points)

        # Generate masks for all points
        # Use negative labels for corner points to help exclude background
        point_labels = np.ones(len(points))
        point_labels[-4:] = 0  # Last 4 points (corners) are negative examples

        masks, scores, _ = self.models[0].predict(point_coords=points, point_labels=point_labels, multimask_output=True)

        # Get the mask with highest score
        best_mask_idx = np.argmax(scores)
        mask = masks[best_mask_idx]

        # Post-process the mask to ensure continuity
        # Convert to uint8 for OpenCV operations
        mask = mask.astype(np.uint8) * 255

        # Find the largest connected component
        num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(mask, connectivity=8)

        if num_labels > 1:  # If we have more than just the background
            # Get the largest component (excluding background)
            largest_label = 1 + np.argmax(stats[1:, cv2.CC_STAT_AREA])
            mask = (labels == largest_label).astype(np.uint8) * 255

        # Clean up the mask with morphological operations
        kernel = np.ones((5, 5), np.uint8)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)  # Close small holes
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)  # Remove small noise

        return mask

    def _load_raw_image(self, file_path, resize_max_dim=1024):
        self.logger.info(f"Attempting to load raw image: {file_path}")
        try:
            self.logger.debug(f"  File exists: {os.path.exists(file_path)}")
            self.logger.debug(f"  File size: {os.path.getsize(file_path) if os.path.exists(file_path) else 'N/A'}")
            self.logger.debug("  Opening file with rawpy...")
            with rawpy.imread(file_path) as raw:
                self.logger.debug("  Post-processing raw image...")
                rgb = raw.postprocess()
            self.logger.debug("  Resizing image...")
            h, w = rgb.shape[:2]
            scale = resize_max_dim / max(h, w)
            resized = cv2.resize(rgb, (int(w * scale), int(h * scale)))
            self.logger.debug("  Successfully loaded and processed image")
            return resized
        except rawpy.LibRawFileUnsupportedError as e:
            self.logger.error(f"  Error: Unsupported RAW format in {os.path.basename(file_path)}: {str(e)}")
            raise
        except rawpy.LibRawIOError as e:
            self.logger.error(f"  Error: IO error reading {os.path.basename(file_path)}: {str(e)}")
            raise
        except Exception as e:
            self.logger.error(f"  Error: Unexpected error processing {os.path.basename(file_path)}: {str(e)}")
            self.logger.error(f"  Error type: {type(e).__name__}")
            self.logger.error(f"  Error details: {str(e)}")
            raise

    def _focus_score(self, image, subject_mask=None):
        """Calculate focus score for both subject area and full image"""
        gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)

        # Calculate full image focus score
        full_image_score = cv2.Laplacian(gray, cv2.CV_64F).var()

        # Calculate subject area focus score if mask is provided
        if subject_mask is not None:
            # Apply mask to focus only on subject
            masked_gray = cv2.bitwise_and(gray, gray, mask=subject_mask.astype(np.uint8))
            # Calculate focus only on non-zero pixels (subject area)
            if np.sum(subject_mask) > 0:
                subject_score = cv2.Laplacian(masked_gray, cv2.CV_64F).var()
            else:
                subject_score = 0
            return subject_score, full_image_score

        return full_image_score, full_image_score

    def _exposure_score(self, image):
        gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
        hist = cv2.calcHist([gray], [0], None, [256], [0, 256])
        hist /= hist.sum()
        underexposed = hist[:30].sum()
        overexposed = hist[-30:].sum()
        return 1 - (underexposed + overexposed)  # Higher is better

    def _halo_score(self, image):
        # Edge detection followed by dilation → compare against original
        gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
        edges = cv2.Canny(gray, 100, 200)
        dilated = cv2.dilate(edges, np.ones((3, 3), np.uint8))
        overlap = cv2.bitwise_and(gray, gray, mask=dilated)
        std_dev = np.std(overlap)
        return std_dev  # Higher could indicate more halos

    def _compute_image_similarity(self, img1, img2):
        """Compute similarity between two images using histogram comparison"""
        # Convert to HSV color space
        hsv1 = cv2.cvtColor(img1, cv2.COLOR_RGB2HSV)
        hsv2 = cv2.cvtColor(img2, cv2.COLOR_RGB2HSV)

        # Calculate histograms
        hist1 = cv2.calcHist([hsv1], [0, 1], None, [8, 8], [0, 180, 0, 256])
        hist2 = cv2.calcHist([hsv2], [0, 1], None, [8, 8], [0, 180, 0, 256])

        # Normalize histograms
        cv2.normalize(hist1, hist1, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)
        cv2.normalize(hist2, hist2, alpha=0, beta=1, norm_type=cv2.NORM_MINMAX)

        # Compare histograms
        similarity = cv2.compareHist(hist1, hist2, cv2.HISTCMP_CORREL)
        return similarity

    def group_similar_photos(self, photos: List[Dict], time_threshold: int = 30) -> List[List[Dict]]:
        """
        Group photos into sequences based on file modification time and content similarity
        time_threshold: maximum seconds between photos to be considered part of same sequence
        """
        # Sort photos by modification time
        photos.sort(key=lambda x: x["mod_time"])

        sequences = []
        current_sequence = []

        for _, photo in enumerate(photos):
            if not current_sequence:
                current_sequence.append(photo)
                continue

            prev_photo = current_sequence[-1]
            time_diff = (photo["mod_time"] - prev_photo["mod_time"]).total_seconds()

            # Check if photos are similar in content and time
            if time_diff <= time_threshold:
                similarity = self._compute_image_similarity(photo["image"], prev_photo["image"])
                if similarity > 0.7:  # Threshold for content similarity
                    current_sequence.append(photo)
                else:
                    if len(current_sequence) > 1:
                        sequences.append(current_sequence)
                    current_sequence = [photo]
            else:
                if len(current_sequence) > 1:
                    sequences.append(current_sequence)
                current_sequence = [photo]

        if len(current_sequence) > 1:
            sequences.append(current_sequence)

        return sequences

    def rank_photos_in_sequence(self, sequence: List[Dict]) -> List[Dict]:
        """Rank photos in a sequence based on their quality scores"""
        # Normalize scores
        max_subject_focus = max(p["subject_focus"] for p in sequence)
        max_full_focus = max(p["full_focus"] for p in sequence)
        max_exposure = max(p["exposure"] for p in sequence)
        min_halo = min(p["halo"] for p in sequence)

        for photo in sequence:
            # Normalize scores to 0-1 range
            photo["normalized_subject_focus"] = (
                photo["subject_focus"] / max_subject_focus if max_subject_focus > 0 else 0
            )
            photo["normalized_full_focus"] = photo["full_focus"] / max_full_focus if max_full_focus > 0 else 0
            photo["normalized_exposure"] = photo["exposure"] / max_exposure if max_exposure > 0 else 0
            photo["normalized_halo"] = 1 - (photo["halo"] / min_halo) if min_halo > 0 else 1

            # Calculate overall score (you can adjust weights)
            photo["overall_score"] = (
                0.4 * photo["normalized_subject_focus"]  # Weight subject focus more heavily
                + 0.2 * photo["normalized_full_focus"]  # Full image focus as secondary
                + 0.3 * photo["normalized_exposure"]  # Exposure importance
                + 0.1 * photo["normalized_halo"]  # Halo has least weight
            )

        # Sort by overall score
        return sorted(sequence, key=lambda x: x["overall_score"], reverse=True)

    def assess_image(self, path):
        """Assess a single image with all models"""
        self.logger.info(f"\nAssessing image: {os.path.basename(path)}")
        try:
            self.logger.debug(f"  Full path: {path}")
            self.logger.debug("  Loading raw image...")
            img = self._load_raw_image(path)
            self.logger.debug("  Getting modification time...")
            mod_time = datetime.fromtimestamp(os.path.getmtime(path))

            self.logger.debug("  Getting subject mask...")
            subject_mask = self._get_subject_mask(img)

            self.logger.debug("  Detecting sky...")
            sky_percentage, sky_mask = self._detect_sky(img, subject_mask)

            self.logger.debug("  Identifying airline...")
            airline_matches = self._identify_airline(img)

            self.logger.debug("  Calculating focus scores...")
            subject_focus, full_focus = self._focus_score(img, subject_mask)

            self.logger.debug("  Calculating exposure and halo scores...")
            exposure = self._exposure_score(img)
            halo = self._halo_score(img)

            self.logger.debug("  Successfully completed assessment")
            return {
                "filename": os.path.basename(path),
                "mod_time": mod_time,
                "image": img,
                "subject_mask": subject_mask,
                "sky_percentage": sky_percentage,
                "sky_mask": sky_mask,
                "airline_matches": airline_matches,
                "subject_focus": subject_focus,
                "full_focus": full_focus,
                "exposure": exposure,
                "halo": halo,
            }
        except Exception as e:
            self.logger.error(f"  Error during assessment of {os.path.basename(path)}: {str(e)}")
            self.logger.error(f"  Error type: {type(e).__name__}")
            self.logger.error(f"  Error details: {str(e)}")
            raise

    def create_assessment_excel(
        self, photos: List[Dict], sequences: List[List[Dict]], output_path: str, directory_path: str
    ) -> None:
        """
        Create or update an Excel file with photo assessments.
        Photos are organized by year in separate tabs.
        New photos are appended to existing sheets.
        Rank 1 photos are highlighted in green.

        Args:
            photos: List of photo assessment dictionaries
            sequences: List of photo sequences
            output_path: Path to the Excel file
            directory_path: Path to the directory being processed (not used for tab names)
        """
        # Load existing workbook or create new one
        if os.path.exists(output_path):
            workbook = openpyxl.load_workbook(output_path)
        else:
            workbook = openpyxl.Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)

        # Create a mapping of photos to their sequence and rank
        photo_to_sequence = {}
        for i, seq in enumerate(sequences, 1):
            ranked_seq = self.rank_photos_in_sequence(seq)
            for j, photo in enumerate(ranked_seq, 1):
                photo_to_sequence[photo["filename"]] = (i, j, photo["overall_score"])

        # Group photos by year
        photos_by_year = {}
        for photo in photos:
            year = photo["mod_time"].strftime("%Y")
            if year not in photos_by_year:
                photos_by_year[year] = []
            photos_by_year[year].append(photo)

        # Process each year's photos
        for year, year_photos in photos_by_year.items():
            sheet_name = f"Photos {year}"

            # Create or get the sheet for this year
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Find the next empty row
                next_row = sheet.max_row + 1
            else:
                sheet = workbook.create_sheet(title=sheet_name)
                next_row = 1
                # Write headers for new sheet
                headers = [
                    "Filename",
                    "Directory",
                    "Date",
                    "Time",
                    "Airline",
                    "Confidence",
                    "Sky %",
                    "Focus Score",
                    "Exposure Score",
                    "Halo Score",
                    "Sequence",
                    "Rank",
                ]
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)
                next_row = 2

            # Write photo data
            row = next_row
            for photo in year_photos:
                # Get sequence and rank info
                seq_info = photo_to_sequence.get(photo["filename"], (0, 0, 0))
                seq_num, rank_num, score = seq_info

                # Create cells for this row
                cells = []
                cells.append(sheet.cell(row=row, column=1, value=os.path.basename(photo["filename"])))
                cells.append(sheet.cell(row=row, column=2, value=os.path.dirname(photo["filename"])))
                cells.append(sheet.cell(row=row, column=3, value=photo["mod_time"].strftime("%Y-%m-%d")))
                cells.append(sheet.cell(row=row, column=4, value=photo["mod_time"].strftime("%H:%M:%S")))

                # Handle airline matches
                if photo["airline_matches"] and photo["airline_matches"][0][1] >= 0.5:
                    airline = photo["airline_matches"][0][0].replace("a ", "").replace("an ", "")
                    confidence = photo["airline_matches"][0][1]
                else:
                    airline = "Unknown"
                    confidence = 0

                cells.append(sheet.cell(row=row, column=5, value=airline))
                cells.append(sheet.cell(row=row, column=6, value=f"{confidence:.1%}"))
                cells.append(sheet.cell(row=row, column=7, value=f"{photo['sky_percentage']:.1%}"))
                cells.append(sheet.cell(row=row, column=8, value=f"{photo['subject_focus']:.2f}"))
                cells.append(sheet.cell(row=row, column=9, value=f"{photo['exposure']:.2f}"))
                cells.append(sheet.cell(row=row, column=10, value=f"{photo['halo']:.2f}"))
                cells.append(
                    sheet.cell(row=row, column=11, value=f"Sequence {seq_num}" if seq_num > 0 else "No Sequence")
                )
                cells.append(
                    sheet.cell(
                        row=row, column=12, value=f"Rank {rank_num} ({score:.2f})" if rank_num > 0 else "No Rank"
                    )
                )

                # Highlight rank 1 photos in green
                if rank_num == 1:
                    green_fill = openpyxl.styles.PatternFill(
                        start_color="92D050", end_color="92D050", fill_type="solid"
                    )
                    for cell in cells:
                        cell.fill = green_fill

                row += 1

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Process photo directories and create assessment Excel file")
    parser.add_argument("--config", required=True, help="Path to the YAML configuration file")
    parser.add_argument("--debug", action="store_true", help="Enable debug mode")
    parser.add_argument("--count-only", action="store_true", help="Only count photos without processing them")
    parser.add_argument(
        "--force", action="store_true", help="Force processing of all directories, even if already processed"
    )
    parser.add_argument(
        "--log-level",
        type=str,
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Set the logging level",
    )
    args = parser.parse_args()

    # Create PhotoJudge instance
    photo_judge = PhotoPicker(
        config_path=args.config,
        debug=args.debug,
        count_only=args.count_only,
        force_reprocess=args.force,
        log_level=args.log_level,
    )

    # Run the photo processing
    photo_judge.run()


if __name__ == "__main__":
    main()
